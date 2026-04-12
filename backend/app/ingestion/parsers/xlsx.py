from pathlib import Path
from openpyxl import load_workbook
from app.ingestion.parsers.base import BaseParser, ParsedPage, ParseResult

class XlsxParser(BaseParser):
    supported_extensions = [".xlsx", ".xls"]

    def parse(self, file_path: Path) -> ParseResult:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        pages = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows.append(" | ".join(cells))
            if rows:
                text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
                pages.append(ParsedPage(text=text, page_number=None, metadata={"sheet": sheet_name}))
        wb.close()
        return ParseResult(pages=pages, metadata={"sheet_count": len(wb.sheetnames)}, file_type="xlsx")
